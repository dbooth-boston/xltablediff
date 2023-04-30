#!/usr/bin/env python3.8

# Copyright 2022 by David Booth
# License: Apache 2.0
# Repo: https://github.com/dbooth-boston/xltablediff
# This code also uses simplediff, https://github.com/paulgb/simplediff/
# which is licensed under the zlib/libpng license:
# <http://www.opensource.org/licenses/zlib-license.php>

# Show value differences between two spreadsheet tables (old and new).
# The tables may also have unrelated leading and/or trailing rows before
# and after the tables.   Those rows are compared as lines, separately from the
# table comparison.
#
# The old and new tables must both have a key column of
# the same name, which is specified by the --key option.  
#
# Optionally, columns from the new table can be merged or appended into the
# old table (using the --merge, --oldAppend or --newAppend options).
#
# Run './xltablediff.py --help' for options and usage info.

# Use these commands for testing:
EXAMPLES = '''
EXAMPLES

# Diff test:
  xltablediff.py  --key ID test1old.xlsx test1new.xlsx --out test1diff.xlsx

# Ignore test:
  xltablediff.py  --key ID --ignore Color test1old.xlsx test1new.xlsx --out test1ignore.xlsx

# Merge test:
  xltablediff.py  --key ID --merge Color test1old.xlsx test1new.xlsx --out test1merge.xlsx

# oldAppend test:
  xltablediff.py  --key ID --oldAppend test1old.xlsx test1new.xlsx --out test1oldAppend.xlsx

# newAppend test:
  xltablediff.py  --key ID --newAppend test1old.xlsx test1new.xlsx --out test1newAppend.xlsx
'''

EXPLANATION = '''
Usually you will need to specify a table key using the '--key ID' option,
where ID is the name of your key column.

A key column must exist in both the old and new table headers.
(Composite keys are not currently supported.)  Keys are used to uniquely
identify the rows in the old and new tables, to determine whether a row
was deleted, added or changed.  The order of rows in each table does
not affect this comparison, because each row is identified by its key,
not by its position in the file.

The first row in a table is the header row, which specifies the names of
the columns.  Body rows follow.  Column names must be unique.  The order
of the columns does not affect their comparison, because each column
is uniquely identified by its column name.  Each cell in the table
is uniquely identified by the combination of its key and column name.
Column names are also used to determine whether an entire column was
deleted or added.

Tables to be compared may appear in any sheet within oldFile and newFile.
If you do not specify which sheet to use, the sheet will be guessed.

A table is not required to begin at the first row of a sheet.  Leading
rows (prior to the table) and trialing rows (after the table) are
permitted -- and will be compared separately from the table comparison
-- but they must not look too much like header rows, or the header row
might not be guessed correctly.  (HINT: Potential header rows contain
only unique non-empty column names, and one of those column names must
match the specified KEY.  Leading and trailing rows must not meet those
criteria.)  The end of the table is taken to be the first row in the
table body that contains an empty key.  Although leading and trailing
rows are permitted, leading and trailing columns are not permitted:
the table must begin in the first column, and no data is permitted after
the last column in the table -- not even in the leading and trailing rows.

Only cell values are compared -- not cell formatting or formulas --
and trailing empty rows or cells are ignored.  If a cell somehow contains
any tabs they will be silently converted to spaces prior to comparison.

The resulting outFile highlights differences found between the
oldFile and newFile tables.  The first column in the outFile
contains a marker indicating whether the row changed:
    -   Row deleted
    +   Row added
    =   Row unchanged (excluding columns added or deleted)
    c-  Row changed: this row shows the old content
    c+  Row changed: this row shows the new content
'''

# Strategy:
# 1. Ignore empty trailing rows and columns.
# 2. Diff the rows before the table only as lines.
# 3. Each row is uniquely identified by the keys in the key column;
# each column is uniquely identified by the headers in the header row.
# 4. Detect added or deleted rows or columns by comparing the old
# and new keys and headers.
# 5. The order of the keys and headers does not matter for diffing.
# It is only used for display purposes.
# 6. Cells are uniquely identify by header and key.
# 7. Detect changed cells by comparing them.
# 8. Use the newFile ordering of headers and keys for
# displaying the results.  If a row or column was deleted, show it
# after the row/column that preceded it in oldFile.  If no
# row/column preceded it in oldFile, then show it before the row/column
# that followed it in oldFile.
# 9. Highlight changes using cell fill colors.

# Ideas for future enhancements: 
# 1. Allow the table to specified (within
# the spreadsheet) by specifying a range, such as: --table=B10:G16
# 2. Optionally suppress unchanged rows and/or columns.

import sys
import os
import os.path
import openpyxl
from openpyxl.styles import PatternFill, Fill, Font
import re
import json
import pprint
import inspect
import keyword
import argparse
import copy

##################### Globals #####################
fillChange =    PatternFill("solid", fgColor="FFFFAA")
fillChangeRow = PatternFill("solid", fgColor="FFFFDD")
fillDelRow =    PatternFill("solid", fgColor="FFB6C1")
fillAddRow =    PatternFill("solid", fgColor="B6FFC1")
fillDelCol =    PatternFill("solid", fgColor="FFDDE2")
# fillAddCol =    PatternFill("solid", fgColor="DDFFE2")
fillAddCol =    PatternFill("solid", fgColor="CCFFC2")
fillKeyCol =    PatternFill("solid", fgColor="E8E8FF")
fillIgnore =    PatternFill("solid", fgColor="E0E0E0")

##################### NoTabs #####################
def NoTabs(rows):
    ''' Change tabs to spaces in all cells.
    Modifies the given rows of values in place.
    '''
    for i in range(len(rows)):
        row = rows[i]
        for j in range(len(row)):
            row[j] = row[j].replace("\t", " ")
    return rows

##################### TrimAndPad #####################
def TrimAndPad(rows):
    ''' Trim trailing empty rows and columns, and pad
    every row to have the same number of values.
    Modifies the given rows of values in place.
    '''
    iLastRow = -1
    jLastColumn = -1
    for i in range(len(rows)):
        row = rows[i]
        # Look for the last non-empty cell in the row:
        jLastThisRow = -1
        for j in range(len(row)-1, -1, -1):
            if row[j]:
                jLastThisRow = j
                if j > jLastColumn:
                    jLastColumn = j
                break
        if jLastThisRow >= 0:
            iLastRow = i
            if jLastThisRow > jLastColumn:
                jLastColumn = jLastThisRow
    nRows = iLastRow + 1
    nColumns = jLastColumn + 1
    del rows[nRows : ]
    # Now pad (or trim) each row to the max number of non-empty columns.
    for i in range(nRows):
        row = rows[i]
        if len(row) > nColumns:
            # Trim a row with extra cells:
            del row[ nColumns : ]
        # Pad a row with too few cells:
        padding = [ "" for j in range(len(row), nColumns) ]
        row.extend(padding)
    return rows

##################### GuessHeaderRow #####################
def GuessHeaderRow(rows, key, title):
    ''' Guess the header row, as the first row with entirely non-empty
    unique cells.  
    If key was specified, the header row also must contain the key.
    Returns:
        iHeaders = the 0-based index of the header row, or None if not found.
        possibleKeys = list of headers with unique non-empty column values
    '''
    iHeaders = None
    headers = None
    nColumns = len(rows[0])
    # sys.stderr.write(f"[INFO] Sheet '{title}' n rows: {len(rows)}\n")
    # r is 0-based index.
    for r in range(len(rows)):
        row = rows[r]
        # sys.stderr.write(f"[INFO] Sheet '{title}' row {r+1}: {repr(row)}\n")
        # sys.stderr.write(f"[INFO] Processing sheet '{title}' row: {r+1}\n")
        nValues = next( (j for j,v in enumerate(row ) if v == ''), len(row) )
        # sys.stderr.write(f"[INFO] nValues: {repr(nValues)}\n")
        if nValues != nColumns: continue
        if key and key not in row: continue
        nSetItems = len(set(row))
        if nSetItems != nValues:
            # Cannot be a header row, because the values are not unique.
            continue
        # Stop at the first qualifying row
        headers = row.copy()
        iHeaders = r
        break
    possibleKeys = []
    if iHeaders is not None:
        # sys.stderr.write(f"headers: {repr(headers)}\n")
        # Collect possibleKeys.
        for j in range(len(headers)):
            colValues = set()
            # Disqualify the column if it contains duplicate keys
            # before the first empty cell (which would indicate the end
            # of the table).
            dupeFound = None
            for i in range(iHeaders+1, len(rows)):
                v = rows[i][j]
                if v == '': break
                if v in colValues:
                    dupeFound = True
                    break
                colValues.add(v)
            if not dupeFound: possibleKeys.append(headers[j])
    # sys.stderr.write(f"possibleKeys: {repr(possibleKeys)}\n")
    return iHeaders, possibleKeys

##################### UniqueName #####################
def UniqueName(header, headerSet):
    ''' Return a new name for header that is unique in headerSet,
    and add the new name to headerSet.  MODIFIES headerSet!
    '''
    if header not in headerSet:
        headerSet.add(header)
        return header
    i = 0
    h = ""
    while True:
        i += 1
        h = header + "_" + str(i)
        if h not in headerSet: break
    headerSet.add(h)
    return h

##################### LoadWorkBook #####################
def LoadWorkBook(file, data_only=True):
    ''' Read a .xlsx file and return the workbook.
    '''
    wb = None
    try:
        sys.stderr.write(f"[INFO] Reading file: '{file}'\n")
        wb = openpyxl.load_workbook(file, data_only) 
    except ValueError as e:
        s = str(e)
        # sys.stderr.write(f"[INFO] Caught exception: '{s}'\n")
        if s.startswith('Value does not match pattern'):
            sys.stderr.write(f"[ERROR] Unable to load file: '{file}'\n If a sheet uses a filter, try eliminating the filter.\n")
            sys.exit(1)
        raise e
    return wb

##################### FindTable #####################
def FindTable(wb, wantedTitle, key, file, maxColumns):
    ''' Read a workbook wb and possibly a wantedTitle, find the desired table.
    Raises an exception if the table is not found.
    Returns a tuple:
        sheet = the sheet in which the table was found.
        rows = rows (values only) of the sheet in which the table was found.
        iHeaders = 0-based index of the header row in rows.
        iTrailing = 0-based index of rows after the table, which begins
            with the first row that lacks a key.
        jKey = Index of key used: either the one that was passed in or the 
            first possibleKey if no key was specified.
    '''
    # Potentially look through all sheets for the one to compare.
    # If the --sheet option was specified, then only that one will be 
    # checked for the desired header.
    sheet = None
    iHeaders = None
    rows = None
    title = ""
    possibleKeys = []
    allPossibleKeys = set()
    for s in wb:
        # sys.stderr.write(f"[INFO] Sheet: '{s.title} type of s: {repr(type(s))}'\n")
        title = s.title.strip()
        if wantedTitle:
            if title == wantedTitle:
                sheet = s
            else:
                # sys.stderr.write(f"[INFO] Skipping unwanted sheet: '{s.title}'\n")
                continue
        TrimSheet(s, maxColumns, file)
        # Get the rows of cells:
        rows = [ [str(Value(v)).strip() for v in valuesRow] for valuesRow in s.values ]
        # sys.stderr.write(f"[INFO] file: '{file}' c.value rows: \n{repr(rows)} \n")
        TrimAndPad(rows)
        NoTabs(rows)
        # Look for the header row.
        iHeaders, possibleKeys = GuessHeaderRow(rows, key, title)
        allPossibleKeys.update(possibleKeys)
        if iHeaders is None:
            # Not found.  Maybe the key is wrong.
            if key:
                # Try again without specifying the key, in order to
                # suggest possible keys.
                iOther, otherPossibleKeys = GuessHeaderRow(rows, None, title)
                allPossibleKeys.update(otherPossibleKeys)
        else:
            # Found a header row.
            sheet = s
            # sys.stderr.write(f"[INFO] In sheet '{s.title}' found table headers at row {iHeaders+1}\n")
        if sheet:
            break
    if not sheet and wantedTitle:
            sys.stderr.write(f"[ERROR] Sheet not found: '{wantedTitle}'\n")
            sys.stderr.write(f" in file: '{file}'\n")
            sys.exit(1)
    if iHeaders is None:
        withKey = ""
        if key: withKey = f" with key '{key}'"
        sys.stderr.write(f"[ERROR] Unable to find header row{withKey}\n")
        sys.stderr.write(f" in file: '{file}'\n")
        sys.stderr.write(f" This can be caused by duplicate column names\n or by having data in a column beyond the table.\n")
        pKeys = " ".join(sorted(map((lambda v: f"'{v}'"), allPossibleKeys)))
        if key and allPossibleKeys:
            sys.stderr.write(f" Potential keys: {pKeys}\n")
        sys.exit(1)
    # Find the key
    originalKey = key
    if key: 
        if key not in rows[iHeaders]:
            sys.stderr.write(f"[ERROR] Key not found in header row {iHeaders+1}: '{key}'\n")
            sys.stderr.write(f" in file: '{file}'  sheet: '{sheet.title}\n")
            sys.exit(1)
        possibleKeys = [key]
        allPossibleKeys = set([key])
    elif not possibleKeys:
        sys.stderr.write(f"[ERROR] No key found in header row {iHeaders+1}\n")
        sys.stderr.write(f" in file: '{file}'  sheet: '{sheet.title}\n")
        sys.exit(1)
    elif len(possibleKeys) == 1: key = possibleKeys[0]
    else:
        # Look for the first header ending with "id":
        key = next( (k for k in possibleKeys if k.lower().endswith('id')), '' )
        # Fall back to taking the first of possibleKeys:
        if not key: key = possibleKeys[0]
    # Find the index of the key:
    jKey = next( (j for j,v in enumerate(rows[iHeaders]) if v == key), -1 )
    assert( jKey >= 0 )
    if not originalKey:
        sys.stderr.write(f"[INFO] Assuming key: '{key}'\n")
    # Find the end of the table: the first row with an empty key (if any).
    iTrailing = next( (i for i in range(iHeaders, len(rows)) if rows[i][jKey] == ''), len(rows) )
    # sys.stderr.write(f"[INFO] jKey: {jKey} file: '{file}'\n")
    # sys.stderr.write(f"[INFO] iHeaders: {iHeaders} file: '{file}'\n")
    # sys.stderr.write(f"[INFO] iTrailing: {iTrailing} file: '{file}'\n")
    # sys.stderr.write(f"[INFO] file: '{file}' rows: \n{repr(rows)} \n")
    return (sheet, rows, iHeaders, iTrailing, jKey)

##################### RemoveTrailingEmpties #####################
def RemoveTrailingEmpties(items):
    ''' Remove trailing empty items from the given list of items,
    returning a new resulting list.
    '''
    nItems = next( (i+1 for i in range(len(items)-1, -1, -1) if items[i]), 0 )
    return items[0 : nItems]

##################### CompareLeadingTrailingRows #####################
def CompareLeadingTrailingRows(diffRows, oldRows, iOldStart, nOldRows, newRows, iNewStart, nNewRows, nDiffHeaders):
    ''' Compare the old and new leading or trailing rows.   Trailing empty cells
    are ignored.  The first item in each returned diffRow is one of {=, -, +}.
    Modifies diffRows in place.
    '''
    # sys.stderr.write(f"CompareLeadingTrailingRows\n")
    # Concatenate the cells in each row, separated by tabs.
    oldLeadingLines = [ "\t".join( RemoveTrailingEmpties(oldRows[i]) ) for i in range(iOldStart, nOldRows) ]
    newLeadingLines = [ "\t".join( RemoveTrailingEmpties(newRows[i]) ) for i in range(iNewStart, nNewRows) ]
    rawDiffs = diff(oldLeadingLines, newLeadingLines)
    # diff returns pairs: (d, dList)
    # where:    
    #   d = diff mark: one of {=, -, +} 
    #   dList = a list of items that were the same, deleted or added.
    # So we need to flatten the dLists, so that we just have pairs: (d, item).
    # flattened = [val for sublist in list_of_lists for val in sublist]
    flatDiffs = [ (d[0], v) for d in rawDiffs for v in d[1] ]
    nLeadingChanges = len([ d for d in flatDiffs if d[0] != '=' ])
    # Prepend the diff mark:
    diffLeadingLines = [ d[0] + "\t" + d[1] for d in flatDiffs ]
    for line in diffLeadingLines:
        partialRow = line.split("\t")
        diffRow = [ (partialRow[i] if i < len(partialRow) else '') for i in range(nDiffHeaders+1) ]
        diffRows.append(diffRow)
    # sys.stderr.write(f"diffRows:\n{repr(diffRows)}\n")
    return nLeadingChanges

##################### CompareHeaders #####################
def CompareHeaders(oldHeaders, oldHeaderIndex, newHeaders, newHeaderIndex):
    ''' Compare the oldHeaders with newHeaders, returning a combined
    list of headers.
    Return:
        diffHeaderMarks = {=, -, +}, one for each diffHeader
        diffHeaders = Combined old and new headers
    '''
    # Headers are treated as column keys: they must be unique.
    # Warn(f"in CompareHeaders oldHeaders: {repr(oldHeaders)}")
    if '' in oldHeaders:
        iEmpty = next( (i for i in range(len(oldHeaders)) if oldHeaders[i] == ''), -1 )
        letter = openpyxl.utils.cell.get_column_letter(iEmpty+1)
        raise ValueError(f"[ERROR] Empty header in column {letter} of old table\n")
    if '' in newHeaders:
        iEmpty = next( (i for i in range(len(newHeaders)) if newHeaders[i] == ''), -1 )
        letter = openpyxl.utils.cell.get_column_letter(iEmpty+1)
        raise ValueError(f"[ERROR] Empty header in column {letter} of new table\n")
    # Build up the diffHeaders from the newHeaders plus any deleted oldHeaders
    diffHeaders = []
    diffHeaderMarks = []    # {=, -, +}
    # First copy any initial deleted headers
    for i, h in enumerate(oldHeaders):
        if h not in newHeaderIndex:
            diffHeaders.append(h)
            diffHeaderMarks.append('-')
        else:
            break
    # Now copy the newHeaders into diffHeader, but each time one has
    # a corresponding oldHeader that has any deleted headers after it,
    # also copy them in.
    for i, h in enumerate(newHeaders):
        diffHeaders.append(h)
        if h in oldHeaderIndex:
            diffHeaderMarks.append('=')
            j = oldHeaderIndex[h] + 1
            while j < len(oldHeaders) and oldHeaders[j] not in newHeaderIndex:
                diffHeaders.append(oldHeaders[j])
                diffHeaderMarks.append('-')
                j += 1
        else:
            diffHeaderMarks.append('+')
    return diffHeaderMarks, diffHeaders

##################### CompareBody #####################
def CompareBody(diffRows, diffHeaders, ignoreHeaders,
        oldRows, oldHeaders, iOldHeaders, iOldTrailing, oldHeaderIndex, jOldKey,
        newRows, newHeaders, iNewHeaders, iNewTrailing, newHeaderIndex, jNewKey):
    ''' Compare rows in the body of the table.  Modifies diffRows
    by appending diffRows for the table body.  The first cell of each diffRow 
    will be one of {=, -, +, c-, c+}.
    '''
    iFirstBodyRow = len(diffRows)
    # Make lists of oldKeys and newKeys.
    oldKeys = [ oldRows[i][jOldKey] for i in range(iOldHeaders+1, iOldTrailing) ]
    # oldKeyIndex will index directly into oldRows, which means that
    # the index is offset by iOldHeaders+1, to get past the leading lines
    # and the header row.
    oldKeyIndex = {}
    for i, v in enumerate(oldKeys):
        r = i+iOldHeaders+1 
        if v == '':
            raise  ValueError(f"[ERROR] Table in oldFile contains an empty key at row {r+1}\n")
        if v in oldKeyIndex:
            raise  ValueError(f"[ERROR] Table in oldFile contains a duplicate key on row {r+1}: '{v}'\n")
        oldKeyIndex[v] = r
    # sys.stderr.write(f"jNewKey: {jNewKey} iNewHeaders: {iNewHeaders} iNewTrailing: {iNewTrailing}\n")
    newKeys = [ newRows[i][jNewKey] for i in range(iNewHeaders+1, iNewTrailing) ]
    # newKeyIndex will index directly into newRows, which means that
    # the index is offset by iNewHeaders+1, to get past the leading lines
    # and the header row.
    newKeyIndex = {}
    for i, v in enumerate(newKeys):
        r = i+iNewHeaders+1 
        if v == '':
            raise  ValueError(f"[INTERNAL ERROR] Table in newFile contains an empty key at row {r+1}\n")
        if v in newKeyIndex:
            raise  ValueError(f"[ERROR] Table in newFile contains a duplicate key on row {r+1}: '{v}'\n")
        newKeyIndex[v] = r
    # Make the diff list of rows.
    # diffRows will not include the header row, but each row in it
    # will have a diff mark {-, +, c-, c+} as its first item.
    # First copy any initial deleted rows
    # sys.stderr.write(f"diffHeaders: {repr(diffHeaders)}\n")
    for k in oldKeys:
        if k not in newKeyIndex:
            oldRow = oldRows[oldKeyIndex[k]]
            oldDiffRow = [ '-' ]
            oldDiffRow.extend( [ (oldRow[oldHeaderIndex[h]] if h in oldHeaderIndex else '') for h in diffHeaders ] )
            diffRows.append(oldDiffRow)
        else:
            break
    # Remove from commonHeaders columns that should be ignored:
    commonHeaders = set(oldHeaderIndex.keys()).intersection(set(newHeaderIndex.keys()))
    # sys.stderr.write(f"commonHeaders: {repr(commonHeaders)}\n")
    ignoreSet = set(ignoreHeaders)
    remainingHeaders = ignoreSet.difference(commonHeaders)
    if remainingHeaders:
        h = sorted(remainingHeaders).join(" ")
        sys.stderr.write(f"[ERROR] Bad --ignore column name(s): h\n Column headers specified with --ignore must exist in both old and new tables\n")
        sys.exit(1)
    compareHeaders = commonHeaders.difference(ignoreSet)
    # Now copy the newRows into diffRows, marking each diff row 
    # as one of {=, -, +, c-, c+}.  Each time a new row has
    # a corresponding old row that has any deleted rows after it,
    # also copy them in.  
    for ii, k in enumerate(newKeys):
        i = ii+iNewHeaders+1
        newRow = newRows[i]
        newDiffRow = [ '+' ]    # This might later become = or c+
        newDiffRowValues = [ (newRow[newHeaderIndex[h]] if h in newHeaderIndex else '') for h in diffHeaders ]
        newDiffRow.extend(newDiffRowValues)
        if k in oldKeyIndex:
            # Key k is in both oldRows and newRows.  
            oldRow = oldRows[oldKeyIndex[k]]
            # Include old values:
            for j, h in enumerate(diffHeaders):
                if h in oldHeaderIndex and h not in newHeaderIndex:
                    newDiffRow[j+1] = oldRow[oldHeaderIndex[h]]
            # Did the rows change (excluding added/deleted columns)?
            isEqual = next( (False for h in compareHeaders if oldRow[oldHeaderIndex[h]] != newRow[newHeaderIndex[h]]), True )
            if isEqual:
                # No values changed in columns that are in common in this row.
                # Only add one row to diffRows.
                newDiffRow[0] = '='     # Change the marker
                diffRows.append(newDiffRow)
            else:
                # Values changed from oldRow to newRow.  
                oldDiffRow = [ 'c-' ]
                for h in diffHeaders:
                    v = ''
                    if h in newHeaderIndex: v = newRow[newHeaderIndex[h]]
                    if h in oldHeaderIndex: v = oldRow[oldHeaderIndex[h]]
                    oldDiffRow.append(v)
                diffRows.append(oldDiffRow)
                newDiffRow[0] = 'c+'     # Change the marker
                diffRows.append(newDiffRow)
            # Also add any following deleted old rows.
            for j in range(oldKeyIndex[k]+1, iOldTrailing):
                if oldKeys[j - (iOldHeaders+1)] in newKeyIndex:
                    break
                oldRow = oldRows[j]
                oldDiffRow = [ '-' ]
                oldDiffRowValues = [ (oldRow[oldHeaderIndex[h]] if h in oldHeaderIndex else '') for h in diffHeaders ]
                oldDiffRow.extend(oldDiffRowValues)
                diffRows.append(oldDiffRow)
            # Finished k in oldKeyIndex
        else:
            # k is not in oldKeyIndex.  k is a new key.
            newDiffRow = [ '+' ]    # This might later become = or c+
            newDiffRowValues = [ (newRow[newHeaderIndex[h]] if h in newHeaderIndex else '') for h in diffHeaders ]
            newDiffRow.extend( newDiffRowValues )
            diffRows.append(newDiffRow)
        # end of loop: for ii, k in enumerate(newKeys):
    # sys.stderr.write(f"diffRows: \n{repr(diffRows)}\n")
    # Count the changes:
    nBodyChanges = 0
    for i in range(iFirstBodyRow, len(diffRows)):
        if diffRows[i][0] in {'+', '-'}: nBodyChanges += 1
        elif diffRows[i][0] == 'c-':
            for j in range(1, len(diffRows[i])):
                if diffRows[i][j] != diffRows[i+1][j]: nBodyChanges += 1
    return nBodyChanges

##################### Info #####################
def Info(s):
    sys.stderr.write(f"[INFO] {s}\n")

##################### Warn #####################
def Warn(s):
    sys.stderr.write(f"[WARNING] {s}\n")

##################### Die #####################
def Die(s):
    sys.stderr.write(f"[ERROR] {s}\n")
    sys.exit(1)

##################### CompareTables #####################
def CompareTables(oldRows, iOldHeaders, iOldTrailing, jOldKey, 
        newRows, iNewHeaders, iNewTrailing, jNewKey, ignoreHeaders, command):
    ''' Compare the old and new tables, and any leading or trailing rows.  
    Returns:
        diffRows = Rows of diff cells.  The first cell of each row is
            a marker with one of: {=, +, -, c-, c+}.
        iDiffHeaders = Index of the first header row.  There will be two
            header rows (old and new) if any headers changed.
        iDiffBody = Index of the table body (either iDiffHeaders+1
            or iDiffHeaders+2).
        iDiffTrailing = Index of any trailing rows after the table, if any.
    '''
    ###### Compare the headers, i.e., the column names.
    # This is done first to figure out how many resulting columns we'll need.
    # The resulting list of headers will be the union of old and new headers,
    # some headers being deleted, some added and some unchanged.
    # The main difficulty is in deciding the best ordering for them,
    # which will basically be the same order as the new headers,
    # but with deleted headers interspersed.
    # Old and new headers are treated as column keys: they must be unique.
    # And they must not contain any empty header.
    oldHeaders = oldRows[iOldHeaders]
    # Warn(f"oldHeaders: {repr(oldHeaders)}")
    oldHeaderIndex = { v: i for i, v in enumerate(oldHeaders) }
    newHeaders = newRows[iNewHeaders]
    # Warn(f"newHeaders: {repr(newHeaders)}")
    newHeaderIndex = { v: i for i, v in enumerate(newHeaders) }
    diffHeaderMarks, diffHeaders = CompareHeaders(oldHeaders, oldHeaderIndex, newHeaders, newHeaderIndex)
    # nDiffHeader does not include the marker column.
    # The total number of columns in diffRows will be nDiffHeaders+1.
    nDiffHeaders = len(diffHeaders)     
    diffRows = []
    ###### Echo the command on the first row?
    if command:
        # Add one to the length for the marker column:
        cmdRow = [ '' for j in range(len(diffHeaders)+1) ]
        cmdRow[0] = '#'
        cmdRow[1] = command
        diffRows.append(cmdRow)
    ###### Compare the lines before the start of the table.
    nChanges = CompareLeadingTrailingRows(diffRows, oldRows, 0, iOldHeaders, newRows, 0, iNewHeaders, nDiffHeaders)
    # Info(f"nLeadingChanges: {nChanges}")
    iDiffHeaders = len(diffRows)
    iDiffBody = iDiffHeaders + 2    # 2 for old and new header rows
    # sys.stderr.write(f"iOldTrailing: {iOldTrailing} iNewTrailing: {iNewTrailing}\n")
    ###### Add old and new headers to diffRows.
    nColumnChanges = 0
    if len(diffHeaders) == len(oldHeaders):
        # No columns were added or deleted.
        iDiffBody = iDiffHeaders + 1    # Only one header row after all
        diffRow = [ '=' ]
        diffRow.extend(oldHeaders)
        diffRows.append(diffRow)
    else:
        # At least one column was added or deleted.
        oldDiffRow = [ 'c-' ]
        oldDiffRow.extend( [ (oldHeaders[oldHeaderIndex[h]] if h in oldHeaderIndex else '') for h in diffHeaders ] )
        diffRows.append(oldDiffRow)
        newDiffRow = [ 'c+' ]
        newDiffRow.extend( [ (newHeaders[newHeaderIndex[h]] if h in newHeaderIndex else '') for h in diffHeaders ] )
        diffRows.append(newDiffRow)
        # Skip the marker column in counting column changes:
        columnChanges = [ j for j in range(1, len(oldDiffRow)) if oldDiffRow[j] != newDiffRow[j] ]
        nColumnChanges = len(columnChanges)
    # Info(f"nColumnChanges: {nColumnChanges}")
    nChanges += nColumnChanges
    ###### Compare the table body rows.
    # Compare rows, excluding columns that were added or deleted.
    nBodyChanges = CompareBody(diffRows, diffHeaders, ignoreHeaders,
        oldRows, oldHeaders, iOldHeaders, iOldTrailing, oldHeaderIndex, jOldKey,
        newRows, newHeaders, iNewHeaders, iNewTrailing, newHeaderIndex, jNewKey)
    # Info(f"nBodyChanges: {nBodyChanges}")
    nChanges += nBodyChanges
    iDiffTrailing = len(diffRows)
    ###### Compare trailing rows (after the table).
    nTrailingChanges = CompareLeadingTrailingRows(diffRows, oldRows, iOldTrailing, len(oldRows), newRows, iNewTrailing, len(newRows), nDiffHeaders)
    # Info(f"nTrailingChanges: {nTrailingChanges}")
    nChanges += nTrailingChanges
    return diffRows, iDiffHeaders, iDiffBody, iDiffTrailing, nChanges

##################### Value #####################
def Value(v):
    ''' Default to the empty string if v is None.
    '''
    return '' if v is None else v

##################### CopyCellAttributes #####################
def CopyCellAttributes(toCell, fromCell):
    ''' Copy openpyxl cell attributes fromCell toCell.
    '''
    # For some unknown reason, 'style' messes up date formats
    # and fills if it is set *after* setting number_format.
    # IDK if it does anything if it is set before, but here it is.
    toCell.style = copy.copy(fromCell.style)
    toCell.fill = copy.copy(fromCell.fill)
    toCell.font = copy.copy(fromCell.font)
    toCell.number_format = copy.copy(fromCell.number_format)

##################### NewAppendTable #####################
def NewAppendTable(oldWb, oldSheet, iOldHeaders, iOldTrailing, jOldKey,
        newSheet, iNewHeaders, iNewTrailing, jNewKey, outFile):
    ''' Append columns of new table to old table.
    newRows take priority: 
    oldRows that do not exist in newRows are discarded, and
    newRows that do not exist in oldRows are added, with
    the key from newRows.
    Write the resulting spreadsheet to outFile.
    '''
    # The strategy is to make a new worksheet having the resulting rows
    # (from newRows) that we want, and then call OldAppendTable to
    # fill it in and write it out.
    outWb = openpyxl.Workbook()
    outWb.iso_dates = True
    outSheet = outWb.active
    oldCellRows = tuple(oldSheet.rows)
    newCellRows = tuple(newSheet.rows)
    nOldRows = len(oldCellRows)
    nNewRows = len(newCellRows)
    oldHeaderCells = oldCellRows[iOldHeaders]   # Tuple
    newHeaderCells = newCellRows[iNewHeaders]   # Tuple
    nOldHeaders = len(oldHeaderCells)
    nNewHeaders = len(newHeaderCells)
    ##### First copy all oldRows (including headers) before the table body
    # iOut is the 0-based row index in outSheet where we'll be writing:
    iOut = 0
    for iOld in range(iOldHeaders+1):
        for jOld in range(nOldHeaders):
            oldCell = oldSheet.cell(iOld+1, jOld+1)
            c = outSheet.cell(iOut+1, jOld+1, copy.copy(oldCell.value))
            CopyCellAttributes(c, oldCell)
        iOut += 1
    iOutHeaders = iOut-1
    assert( iOutHeaders == iOldHeaders )
    ##### Make a new row for any initial newRows that are not in oldRows:
    oldKeyIndex = { oldSheet.cell(i+1, jOldKey+1).value: i for i in range(iOldHeaders+1, iOldTrailing) }
    newKeyIndex = { newSheet.cell(i+1, jNewKey+1).value: i for i in range(iNewHeaders+1, iNewTrailing) }
    for iNew in range(iNewHeaders+1, iNewTrailing):
        kNew = newSheet.cell(iNew+1, jNewKey+1).value
        if kNew in oldKeyIndex: break
        # sys.stderr.write(f"Making row for kNew: {repr(kNew)}\n")
        # kNew is not in oldRows.  Make a new row for it.
        for jOld in range(nOldHeaders):
            v = ''
            if jOld == jOldKey: v = kNew
            c = outSheet.cell(iOut+1, jOld+1, v)
            c.fill = fillAddRow
        iOut += 1
    ##### Found the first shared row.  
    # Now, starting with the first oldRow that is also in newRows,
    # copy oldRows that are also in newRows, but keep them in the
    # oldRows order.
    for iOld in range(iOldHeaders+1, iOldTrailing):
        kOld = oldSheet.cell(iOld+1, jOldKey+1).value
        # Ignore oldRows that are not in newRows:
        if kOld not in newKeyIndex: continue
        # This row is in both.  Copy it.
        for jOld in range(nOldHeaders):
            oldCell = oldSheet.cell(iOld+1, jOld+1)
            c = outSheet.cell(iOut+1, jOld+1, copy.copy(oldCell.value))
            CopyCellAttributes(c, oldCell)
        iOut += 1
        # Now make a new row for each following newRow that is not in oldRows.
        for iNew in range(newKeyIndex[kOld]+1, iNewTrailing):
            kNew = newSheet.cell(iNew+1, jNewKey+1).value
            if kNew in oldKeyIndex: break
            for jOld in range(nOldHeaders):
                v = ''
                if jOld == jOldKey: v = kNew
                c = outSheet.cell(iOut+1, jOld+1, v)
                c.fill = fillAddRow
            iOut += 1
    iOutTrailing = iOut
    ##### Copy any trailing oldRows.
    for iOld in range(iOldTrailing, nOldRows):
        for jOld in range(nOldHeaders):
            oldCell = oldSheet.cell(iOld+1, jOld+1)
            c = outSheet.cell(iOut+1, jOld+1, copy.copy(oldCell.value))
            CopyCellAttributes(c, oldCell)
        iOut += 1

    assert( iOut == nOldRows - (iOldTrailing - iOldHeaders) + (iNewTrailing - iNewHeaders) )
    ##### Now can can use OldAppendTable to append the newRows, 
    # because we know that the tables are aligned.
    OldAppendTable(outWb, outSheet, iOutHeaders, iOutTrailing, jOldKey,
        newSheet, iNewHeaders, iNewTrailing, jNewKey, outFile)

##################### OldAppendTable #####################
def OldAppendTable(oldWb, oldSheet, iOldHeaders, iOldTrailing, jOldKey,
        newSheet, iNewHeaders, iNewTrailing, jNewKey, outFile):
    ''' Append columns of new table to old table.
    If priorityNew is True, then newRows take priority: 
    oldRows that do not exist in newRows are discarded, and
    newRows that do not exist in oldRows are added, with
    the key from newRows.
    Write the resulting spreadsheet to outFile.
    oldWb (and oldSheet) are modified in memory!
    '''
    # Get the old and new headers
    oldCellRows = tuple(oldSheet.rows)
    newCellRows = tuple(newSheet.rows)
    oldHeaderCells = oldCellRows[iOldHeaders]   # Tuple
    newHeaderCells = newCellRows[iNewHeaders]   # Tuple
    nOldHeaders = len(oldHeaderCells)
    nNewHeaders = len(newHeaderCells)
    # Extend oldSheet with more columns (for newRows):
    # sys.stderr.write(f"[INFO] nOldHeaders: '{repr(nOldHeaders)} nNewHeaders: '{repr(nNewHeaders)}'\n")
    oldSheet.insert_cols(nOldHeaders+1, nNewHeaders)
    d = oldSheet.calculate_dimension()
    nTotalColumns = nOldHeaders + nNewHeaders
    # Set this again, in case they changed from adding columns:
    oldCellRows = tuple(oldSheet.rows)
    oldHeaderCells = oldCellRows[iOldHeaders]   # Tuple
    # Get the old and new key columns
    oldCellColumns = tuple(oldSheet.columns)
    newCellColumns = tuple(newSheet.columns)
    oldKeyColumn = oldCellColumns[jOldKey]
    newKeyColumn = newCellColumns[jNewKey]
    # Make a newKeyIndex, to look up new rows from old keys
    newKeyIndex = { newKeyColumn[i].value: i for i in range(iNewHeaders+1, iNewTrailing) }
    # Copy existing newRows into oldRows
    emptyNewCellRowValues = [ None for j in range(nNewHeaders) ]
    emptyNewCellRowFills = [ fillAddRow for j in range(nNewHeaders) ]
    headerSet = set([ c.value for c in oldHeaderCells ])
    for i in range(iOldHeaders, iOldTrailing):
        newCellRowValues = emptyNewCellRowValues
        newCellRowFills = emptyNewCellRowFills
        k = oldKeyColumn[i].value
        assert k is not None
        if i == iOldHeaders:
            # Header row
            newCellRowValues = [ UniqueName(c.value, headerSet) for c in newHeaderCells ]
            newCellRowFills  = [ copy.copy(c.fill)  for c in newHeaderCells ]
        elif k in newKeyIndex:
            # Row is in both old and new
            newCellRow = newCellRows[newKeyIndex[k]]
            newCellRowValues = [ newCellRow[j].value for j in range(nNewHeaders) ]
            newCellRowFills  = [ copy.copy(newCellRow[j].fill) for j in range(nNewHeaders) ]
        assert len(newCellRowValues) == len(newCellRowFills)
        assert len(newCellRowValues) == nNewHeaders
        # sys.stderr.write(f"[INFO] newCellRow: '{repr(newCellRow)}'\n")
        # sys.stderr.write(f"[DEBUG] len newCellRowFills: '{len(newCellRowFills)} newCellRowFills: '{repr(newCellRowFills)}'\n")
        # Copy new cells into the result:
        for j, v in enumerate(newCellRowValues):
            c = oldSheet.cell(i+1, nOldHeaders+j+1, v)
            # sys.stderr.write(f"[DEBUG] c: '{repr(c)} j: '{repr(j)} newCellRowFills[j]: '{repr(newCellRowFills[j])}'\n")
            c.fill = newCellRowFills[j]
            # c.fill = fillAddRow
            # sys.stderr.write(f"[INFO] nOldHeaders: '{repr(nOldHeaders)} j: '{repr(j)}'\n")
    # Write the output file
    # oldSheet.title += '-Appended'
    oldWb.save(outFile)
    sys.stderr.write(f"[INFO] Wrote: '{outFile}'\n\n")

##################### MergeTable #####################
def MergeTable(oldWb, oldSheet, oldRows, iOldHeaders, iOldTrailing, jOldKey,
        newSheet, newRows, iNewHeaders, iNewTrailing, jNewKey, outFile, mergeHeaders):
    ''' Merge specified columns of new table to old table,
    modifying oldWb/oldSheet in place (in memory).
    Write the resulting spreadsheet to outFile.
    '''
    oldHeaders = oldRows[iOldHeaders]
    newHeaders = newRows[iNewHeaders]
    # sys.stderr.write(f"[INFO] MergeTable n columns in oldHeaders: '{len(oldHeaders)}'\n")
    # sys.stderr.write(f"[INFO] MergeTable n columns in oldSheet: '{len(list(oldSheet.rows)[0])}'\n")
    # Prepare to highlight the new columns.
    oldWsRows = tuple(oldSheet.rows)
    newWsRows = tuple(newSheet.rows)
    # newKeyIndex will be used to look up rows in newRows.
    newKeyIndex = {}
    for i in range(iNewHeaders+1, iNewTrailing):
        v = newRows[i][jNewKey]
        if v == '':
            raise  ValueError(f"[ERROR] Table in newFile contains an empty key at row {i+1}\n")
        if v in newKeyIndex:
            raise  ValueError(f"[ERROR] Table in newFile contains a duplicate key on row {i+1}: '{v}'\n")
        newKeyIndex[v] = i
    newHeaderIndex = { h: j for j, h in enumerate(newHeaders) }
    # Looping through columns first (instead of rows) because we can skip
    # the entire column if it is not in mergeHeaders.
    for jOld in range(len(oldHeaders)):
        h = oldHeaders[jOld]
        if h not in mergeHeaders: 
            continue
        jNew = newHeaderIndex[h]
        for i in range(iOldHeaders+1, iOldTrailing):
            oldRow = oldRows[i]
            # oldRow has the string value.  Instead, compare the original
            # value (with its original type).
            oldValue = Value(oldWsRows[i][jOld].value)
            oldKey = oldRow[jOldKey]
            if oldKey in newKeyIndex:
                newValue = Value(newWsRows[newKeyIndex[oldKey]][jNew].value)
                if newValue != oldValue:
                    oldWsRows[i][jOld].value = newValue
                    if oldValue is None or oldValue == '':
                        # New value
                        oldWsRows[i][jOld].fill = fillAddCol
                    else:
                        # Value changed
                        oldWsRows[i][jOld].fill = fillChange
    # Write the output file
    # oldSheet.title += '-Merged'
    oldWb.save(outFile)
    sys.stderr.write(f"[INFO] Wrote: '{outFile}'\n")

##################### FirstNonEmpty #####################
def FirstNonEmpty(cellList):
    ''' Return the 0-based index of the first non-empty cell, or -1.
    Only cell values are examined, as strings.
    '''
    iUsed = next( (i for i in range(len(cellList)) if str(Value(cellList[i].value)).strip()  != ''), -1)
    return iUsed

##################### TrimSheet #####################
def TrimSheet(sheet, maxColumns, filename):
    ''' Modifies the sheet in place, by trimming empty trailing 
    rows and columns.
    '''
    rows = list(sheet.rows)
    nRows = len(rows)
    columns = list(sheet.columns)
    nColumns = len(columns)
    oldNColumns = nColumns
    oldNRows = nRows
    if maxColumns and maxColumns < sheet.max_column:
        sys.stderr.write(f"[INFO] File '{filename}' sheet '{sheet.title}': Deleting {sheet.max_column-maxColumns} columns due to '--maxColumns={maxColumns}'\n")
        # Warn if non-empty columns are found in the next two
        # columns after maxColumns:
        for j in range(maxColumns, maxColumns+2):
            if j >= sheet.max_column: break
            column = columns[j]
            # iUsed = next( (i for i in range(nRows) if str(Value(column[i].value)).strip()  != ''), -1)
            iUsed = FirstNonEmpty(column)
            if iUsed >= 0:
                letter = openpyxl.utils.cell.get_column_letter(iUsed+1)
                sys.stderr.write(f"[WARNING] File '{filename}' sheet '{sheet.title}': Deleting column {j+1} ({letter}) with non-empty data\n")
                break
        sheet.delete_cols(maxColumns, sheet.max_column-maxColumns)
        # Get these again, in case they changed:
        columns = list(sheet.columns)
        nColumns = len(columns)
        rows = list(sheet.rows)
        nRows = len(rows)
    if nColumns > 100 and not maxColumns:
        sys.stderr.write(f"[WARNING] File '{filename}' sheet '{sheet.title}' has a large number of columns: {nColumns}.\n Trimming empty trailing columns may take a long time\n If you are certain that no more than N columns are used in any sheet, you can\n specify the '--maxColumns=N' option (where N is an integer) to delete\n all extra columns.")
    # sys.stderr.write(f"Trimming empty rows and columns. oldNRows: {oldNRows} oldNColumns: {oldNColumns} nRows: {nRows} nColumns: {nColumns} ...\n")
    try:
        # Delete empty trailing rows:
        while nRows > 0:
            row = rows[nRows-1]
            # jUsed = next( (j for j in range(nColumns) if str(Value(row[j].value)).strip()  != ''), -1)
            jUsed = FirstNonEmpty(row)
            if jUsed >= 0: break
            # sys.stderr.write(f"Trimming empty row: {nRows}\n")
            sheet.delete_rows(nRows)
            nRows -= 1
        if nRows == 0: nColumns = 0
    except AttributeError as e:
       raise AttributeError(str(e) + f"\n at sheet '{sheet.title}' row {nRows}")
    try:
        # Delete empty trailing columns:
        warned = False
        while nColumns > 0:
            column = columns[nColumns-1]
            # iUsed = next( (i for i in range(nRows) if str(Value(column[i].value)).strip()  != ''), -1)
            iUsed = FirstNonEmpty(column)
            if iUsed >= 0: break
            # sys.stderr.write(f"Trimming empty column: {nColumns}\n")
            sheet.delete_cols(nColumns)
            nColumns -= 1
            if oldNColumns-nColumns >= 10 and nColumns >= 100 and not warned:
                sys.stderr.write(f"[WARNING] '{filename}' sheet '{sheet.title}': Trimming empty trailing columns from {oldNColumns} columns.\n If this takes too long, consider the '--maxColumns=N' option ...\n")
                warned = True
        if nRows != oldNRows or nColumns != oldNColumns:
            sys.stderr.write(f"[INFO] File '{filename}' sheet '{sheet.title}': Trimmed {oldNRows-nRows} empty trailing rows and {oldNColumns-nColumns} columns\n")
    except AttributeError as e:
       raise AttributeError(str(e) + f"\n at sheet '{sheet.title}' column {nColumns}")
    # sys.stderr.write(f"[INFO] Copying header len: '{repr(len(newCellRows[iNewHeaders]))} nNewHeaders: '{repr(nNewHeaders)}'\n newCellValue Headers: {repr(newCellValues)}\n newHeaders: {repr(newHeaders)}\n")

##################### WriteDiffFile #####################
def WriteDiffFile(diffRows, iDiffHeaders, iDiffBody, iDiffTrailing, oldKey, ignoreHeaders, outFile):
    ''' Write the diffs to the outFile as XLSX, highlighting
    changed rows/columns/cells.
    '''
    nColumns = len(diffRows[0])     # Includes marker column
    oldHeaders = diffRows[iDiffHeaders]
    newHeaders = diffRows[iDiffHeaders]
    if iDiffBody > iDiffHeaders + 1:
        # At least one header changed from old to new, so we'll have
        # two header rows instead of one.
        newHeaders = diffRows[iDiffHeaders+1]
    jKey = next( j for j in range(nColumns) if oldHeaders[j] == oldKey )
    # Create the Excel spreadsheet and fill it with data.
    outWb = openpyxl.Workbook()
    outWb.iso_dates = True
    outSheet = outWb.active
    # Fill the sheet with data
    for diffRow in diffRows:
        outSheet.append(diffRow)
    # Make a new workbook and copy the table into it.

    # Determine column highlights for added/deleted columns. 
    # colFills will highlight added or deleted columns.
    colFills = [ None for j in range(nColumns) ]
    ignoreSet = set(ignoreHeaders)
    for j in range(nColumns):
        if oldHeaders[j] == '':  colFills[j] = fillAddCol
        if newHeaders[j] == '':  colFills[j] = fillDelCol
        # It's enough to check oldHeaders hear, because ignoreSet
        # only includes headers that are in both old and new:
        if oldHeaders[j] in ignoreSet: colFills[j] = fillIgnore
    # Highlight the spreadsheet.
    wsRows = tuple(outSheet.rows)
    for i, diffRow in enumerate(diffRows):
        rowMark = diffRow[0]
        rowFill = None
        if rowMark == '-': rowFill = fillDelRow 
        if rowMark == '+': rowFill = fillAddRow 
        if rowMark == 'c-' or rowMark == 'c+': rowFill = fillChangeRow 
        if i < iDiffHeaders or i >= iDiffTrailing:
            # This is a leading or trailing row.
            # Only use row fills.
            if rowFill:
                for j in range(nColumns):
                    wsRows[i][j].fill = rowFill
        else:
            # This is either a header row or a body row.
            # Apply row fills first, so they'll be overridden
            # by column fills.
            for j in range(nColumns):
                if oldHeaders[j] == oldKey: 
                    wsRows[i][j].fill = fillKeyCol
                if rowMark and rowFill:
                    wsRows[i][j].fill = rowFill
                if i < iDiffBody:
                    # A header row
                    wsRows[i][j].fill = fillKeyCol
                if colFills[j]:
                    wsRows[i][j].fill = colFills[j]
                if j>0 and rowMark == 'c+' and (not colFills[j]) and diffRow[j] != diffRows[i-1][j]:
                    # Highlight this cell and the one above it, if non-empty
                    if diffRows[i-1][j]: wsRows[i-1][j].fill = fillDelRow
                    if diffRows[i][j]:   wsRows[i][j].fill =   fillAddRow
    outSheet.title = 'Differences'
    outWb.save(outFile)
    sys.stderr.write(f"[INFO] Wrote: '{outFile}'\n")

##################### main #####################
def main():
    # Parse command line options:
    global EXPLANATION
    global EXAMPLES
    argParser = argparse.ArgumentParser(
        description='Compare tables in two .xlsx spreadsheets', 
        epilog=EXPLANATION+EXAMPLES,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    argParser.add_argument('--key',
                    help='Specifies KEY as the name of the key column, i.e., its header.  If KEY is of the form "OLDKEY=NEWKEY" then OLDKEY and NEWKEY are the corresponding key columns of the old and new tables, respectively.')
    argParser.add_argument('--oldSheet',
                    help='Specifies the sheet in oldFile to be compared.  Default: the first sheet with a table with a KEY column.')
    argParser.add_argument('--newSheet',
                    help='Specifies the sheet in newFile to be compared.  Default: the first sheet with a table with a KEY column.')
    argParser.add_argument('--sheet',
                    help='Specifies the sheet to be compared, in both oldFile and newFile.  Default: the first sheet with a table with a KEY column.')
    argParser.add_argument('--ignore', nargs=1, action='extend',
                    help='Ignore the specified column when comparing old and new table rows.  This option may be repeated to ignore multiple columns.  The specified column must exist in both old and new tables.')

    argParser.add_argument('--oldAppend', action='store_true',
                    help='''Copy the values of oldFile sheet, appending columns of newFile, 
keeping only the rows of oldFile.  Rows of newFile that do not exist in
oldFile are discarded.  Leading and trailing rows of newFile (before
and after the table) are also discarded.  The number of rows in the
output file will be the same is in oldFile.''')

    argParser.add_argument('--newAppend', action='store_true',
                    help='''Copy the values of oldFile sheet, appending columns of newFile, 
but forcing the resulting table body to have the same rows as in
newFile, based on the keys in newFile.  Rows of newFile that do not
exist in oldFile are inserted (with the key value from newFile), and
rows of oldFile that do not exist in newFile are discarded.  Leading and
trailing rows of newFile (before and after the table) are also discarded.
The number of rows in the resulting table body will be the same is
in newFile.''')

    argParser.add_argument('--merge', nargs=1, action='extend',
                    help='Output a copy of the old sheet, with values from the specifed MERGE column from the new table merged in.  This option may be repeated to merge more than one column.  The number of rows in the output file will be the same is in oldFile.')
    argParser.add_argument('--mergeAll', action='store_true',
                    help="Same as '--merge C' for all non-key columns C that are in both the old and new tables.")
    argParser.add_argument('--maxColumns', type=int,
                    help='Delete all columns after column N, where N is an integer (origin 1).  0 means no limit.  Default: 100.')
    argParser.add_argument('oldFile', metavar='oldFile.xlsx', type=str,
                    help='Old spreadsheet (*.xlsx)')
    argParser.add_argument('newFile', metavar='newFile.xlsx', type=str,
                    help='New spreadsheet (*.xlsx)')
    argParser.add_argument('--out',
                    help='Output file of differences.  This "option" is actually REQUIRED.', required=True)
    # sys.stderr.write(f"[INFO] calling print_using....\n")
    args = argParser.parse_args()
    if args.sheet and args.oldSheet:
        raise ValueError(f"[ERROR] Illegal combination of options: --sheet with --oldSheet")
    if args.sheet and args.newSheet:
        raise ValueError(f"[ERROR] Illegal combination of options: --sheet with --newSheet")
    oldSheetTitle = args.sheet
    newSheetTitle = args.sheet
    if args.oldSheet:
        oldSheetTitle = args.oldSheet
    if args.newSheet:
        newSheetTitle = args.newSheet
    oldKey = None
    newKey = None
    if args.key:
        oldNewKeys = [ k.strip() for k in args.key.split("=") ]
        if len(oldNewKeys) > 2:
            sys.stderr.write(f"[ERROR] Too many keys specified: '--key {args.key}'\n")
            sys.exit(1)
        oldKey = oldNewKeys[0]
        newKey = oldNewKeys[0]
        if len(oldNewKeys) > 1:
            newKey = oldNewKeys[1]
        if oldKey == '' or newKey == '':
            sys.stderr.write(f"[ERROR] Key must not be empty: '--key {args.key}'\n")
            sys.exit(1)
    outFile = args.out
    if not outFile:
        sys.stderr.write("[ERROR] Output filename must be specified: --out=outFile.xlsx\n")
        sys.stderr.write(Usage())
        sys.exit(1)
    if outFile == args.oldFile or outFile == args.newFile:
        sys.stderr.write(f"[ERROR] Output filename must differ from newFile and oldFile: {outFile}\n")
        sys.exit(1)
    maxColumns = -1
    if args.maxColumns: maxColumns = args.maxColumns
    if maxColumns == -1: maxColumns = 100
    # sys.stderr.write("args: \n" + repr(args) + "\n\n")
    oldWb = LoadWorkBook(args.oldFile, data_only=False)
    newWb = LoadWorkBook(args.newFile, data_only=False)
    ####### Determine sheets to compare
    oldSheetTitles = [ s.title for s in oldWb if (not oldSheetTitle) or s.title == oldSheetTitle ]
    newSheetTitles = [ s.title for s in newWb if (not newSheetTitle) or s.title == newSheetTitle ]
    if not oldSheetTitles:
        raise ValueError(f"[ERROR] Sheet '{oldSheetTitle}' not found in oldFile: '{args.oldFile}'")
    if not newSheetTitles:
        raise ValueError(f"[ERROR] Sheet '{newSheetTitle}' not found in newFile: '{args.newFile}'")
    # Default to all matching sheet titles:
    newTitleSet = set(newSheetTitles)
    titlePairs = [ (t, t) for t in oldSheetTitles if t in newTitleSet ]
    if len(oldSheetTitles) == 1: oldSheetTitle = oldSheetTitles[0]
    if len(newSheetTitles) == 1: newSheetTitle = newSheetTitles[0]
    if oldSheetTitle and not newSheetTitle: newSheetTitle = newSheetTitles[0]
    if newSheetTitle and not oldSheetTitle: oldSheetTitle = oldSheetTitles[0]
    # Single sheet comparision:
    if newSheetTitle: titlePairs = [ (oldSheetTitle, newSheetTitle) ]
    #######
    # STOPPED HERE.  I was starting to make this work on multiple sheets.
    # I got as far as constructing titlePairs to be the pairs of sheets to
    # compare (though not tested), but refactoring is needed to: 1. move
    # the output workbook creation and file writing where it can be outside
    # a loop that loops through the titlePairs; and 2. also move the
    # argument processing where it can be outside that new loop.
    # New loop would like this:
    #     for oldSheetTitle, newSheetTitle in titlePairs:
    #######
    # These will be rows of values-only:
    (oldSheet, oldRows, iOldHeaders, iOldTrailing, jOldKey) = FindTable(oldWb, oldSheetTitle, oldKey, args.oldFile, maxColumns)
    oldTitle = oldSheet.title
    if iOldHeaders is None:
        raise ValueError(f"[ERROR] Could not find header row in sheet '{oldSheetTitle}' in oldFile: '{args.oldFile}'")
    sys.stderr.write(f"[INFO] In '{args.oldFile}' sheet '{oldTitle}' found table in rows {iOldHeaders+1}-{iOldTrailing} columns 1-{len(oldRows[0])}\n")
    (newSheet, newRows, iNewHeaders, iNewTrailing, jNewKey) = FindTable(newWb, newSheetTitle, newKey, args.newFile, maxColumns)
    newTitle = newSheet.title
    if iNewHeaders is None:
        raise ValueError(f"[ERROR] Could not find header row in sheet '{newTitle}' in newFile: '{args.newFile}'")
    # Disable the ability to compare trailing rows, because this was found
    # to be error prone: if the table contained a blank row (or key), the 
    # rest of the table would be treated as trailing rows instead of being
    # included in the table comparison.
    if iOldTrailing < len(oldRows):
        Die(f"Trailing non-table rows found at row {iOldTrailing+1}\n"
            + f" in sheet '{oldTitle}' in oldFile: '{args.oldFile}'"
            + f" Trailing rows are now disallowed because they were\n"
            + f" found to be error prone.")
    if iNewTrailing < len(newRows):
        Die(f"Trailing non-table rows found at row {iNewTrailing+1}\n"
            + f" in sheet '{newTitle}' in newFile: '{args.newFile}'"
            + f" Trailing rows are now disallowed because they were\n"
            + f" found to be error prone.")
    sys.stderr.write(f"[INFO] In '{args.newFile}' sheet '{newTitle}' found table in rows {iNewHeaders+1}-{iNewTrailing} columns 1-{len(newRows[0])}\n")
    if args.oldAppend and (args.merge or args.mergeAll):
        sys.stderr.write(f"[ERROR] Options --oldAppend and --merge cannot be used together.\n")
        sys.exit(1)
    if args.newAppend and (args.merge or args.mergeAll):
        sys.stderr.write(f"[ERROR] Options --newAppend and --merge cannot be used together.\n")
        sys.exit(1)
    if args.newAppend and args.oldAppend:
        sys.stderr.write(f"[ERROR] Options --newAppend and --oldAppend cannot be used together.\n")
        sys.exit(1)
    if args.mergeAll or args.merge:
        if args.mergeAll and args.merge:
            sys.stderr.write(f"[ERROR] Options --mergeAll and --merge cannot be used together\n")
            sys.exit(1)
        # sys.stderr.write(f"args.merge: {repr(args.merge)}\n")
        oldHeadersSet = set(oldRows[iOldHeaders])
        newHeadersSet = set(newRows[iNewHeaders])
        mergeHeaders = set()
        if args.mergeAll:
            oldNonKeys = oldHeadersSet.difference(set([oldKey]))
            newNonKeys = newHeadersSet.difference(set([newKey]))
            mergeHeaders = oldNonKeys.intersection(newNonKeys)
        else:
            for h in args.merge:
                # sys.stderr.write(f"h: {repr(h)}\n")
                if h == oldKey or h == newKey:
                    sys.stderr.write(f"[ERROR] Key columns cannot be merged: --merge={'h'}\n")
                    sys.exit(1)
                if (h not in oldHeadersSet):
                    sys.stderr.write(f"[ERROR] Column specified in --merge='{h}' does not exist in old table.\n")
                    sys.exit(1)
                if (h not in newHeadersSet):
                    sys.stderr.write(f"[ERROR] Column specified in --merge='{h}' does not exist in new table.\n")
                    sys.exit(1)
                mergeHeaders.add(h)
        MergeTable(oldWb, oldSheet, oldRows, iOldHeaders, iOldTrailing, jOldKey,
            newSheet, newRows, iNewHeaders, iNewTrailing, jNewKey, outFile, mergeHeaders)
        sys.exit(0)
    if args.oldAppend:
        OldAppendTable(oldWb, oldSheet, iOldHeaders, iOldTrailing, jOldKey,
            newSheet, iNewHeaders, iNewTrailing, jNewKey, outFile)
        sys.exit(0)
    if args.newAppend:
        NewAppendTable(oldWb, oldSheet, iOldHeaders, iOldTrailing, jOldKey,
            newSheet, iNewHeaders, iNewTrailing, jNewKey, outFile)
        sys.exit(0)

    ignoreHeaders = args.ignore if args.ignore else []
    # command will be the command string to echo in the first row output.
    command = ''
    argv = sys.argv.copy()
    argv[0] = os.path.basename(argv[0])
    command = " ".join(argv)
    diffRows, iDiffHeaders, iDiffBody, iDiffTrailing, nChanges = CompareTables(oldRows, iOldHeaders, iOldTrailing, jOldKey,
        newRows, iNewHeaders, iNewTrailing, jNewKey, ignoreHeaders, command)
    Info(f"{nChanges} total differences found")
    oldKey = oldRows[iOldHeaders][jOldKey]
    WriteDiffFile(diffRows, iDiffHeaders, iDiffBody, iDiffTrailing, oldKey, ignoreHeaders, outFile)
    if nChanges == 0: sys.exit(0)
    else: sys.exit(1)


############################################################################
############################## simplediff ##################################
############################################################################
# Lines below here are from the python version of simplediff
# https://github.com/paulgb/simplediff/
# and used under its specified license:
'''
Copyright (c) 2008 - 2013 Paul Butler and contributors

This sofware may be used under a zlib/libpng-style license:

This software is provided 'as-is', without any express or implied warranty. In
no event will the authors be held liable for any damages arising from the use
of this software.

Permission is granted to anyone to use this software for any purpose, including
commercial applications, and to alter it and redistribute it freely, subject to
the following restrictions:

1. The origin of this software must not be misrepresented; you must not claim
that you wrote the original software. If you use this software in a product, an
acknowledgment in the product documentation would be appreciated but is not
required.

2. Altered source versions must be plainly marked as such, and must not be
misrepresented as being the original software.

3. This notice may not be removed or altered from any source distribution.
'''

'''
Simple Diff for Python version 1.0

Annotate two versions of a list with the values that have been
changed between the versions, similar to unix's `diff` but with
a dead-simple Python interface.

(C) Paul Butler 2008-2012 <http://www.paulbutler.org/>
May be used and distributed under the zlib/libpng license
<http://www.opensource.org/licenses/zlib-license.php>
'''

__all__ = ['diff', 'string_diff', 'html_diff']
__version__ = '1.0'


def diff(old, new):
    '''
    Find the differences between two lists. Returns a list of pairs, where the
    first value is in ['+','-','='] and represents an insertion, deletion, or
    no change for that list. The second value of the pair is the list
    of elements.

    Params:
        old     the old list of immutable, comparable values (ie. a list
                of strings)
        new     the new list of immutable, comparable values
   
    Returns:
        A list of pairs, with the first part of the pair being one of three
        strings ('-', '+', '=') and the second part being a list of values from
        the original old and/or new lists. The first part of the pair
        corresponds to whether the list of values is a deletion, insertion, or
        unchanged, respectively.

    Examples:
        >>> diff([1,2,3,4],[1,3,4])
        [('=', [1]), ('-', [2]), ('=', [3, 4])]

        >>> diff([1,2,3,4],[2,3,4,1])
        [('-', [1]), ('=', [2, 3, 4]), ('+', [1])]

        >>> diff('The quick brown fox jumps over the lazy dog'.split(),
        ...      'The slow blue cheese drips over the lazy carrot'.split())
        ... # doctest: +NORMALIZE_WHITESPACE
        [('=', ['The']),
         ('-', ['quick', 'brown', 'fox', 'jumps']),
         ('+', ['slow', 'blue', 'cheese', 'drips']),
         ('=', ['over', 'the', 'lazy']),
         ('-', ['dog']),
         ('+', ['carrot'])]

    '''

    # Create a map from old values to their indices
    old_index_map = dict()
    for i, val in enumerate(old):
        old_index_map.setdefault(val,list()).append(i)

    # Find the largest substring common to old and new.
    # We use a dynamic programming approach here.
    # 
    # We iterate over each value in the `new` list, calling the
    # index `inew`. At each iteration, `overlap[i]` is the
    # length of the largest suffix of `old[:i]` equal to a suffix
    # of `new[:inew]` (or unset when `old[i]` != `new[inew]`).
    #
    # At each stage of iteration, the new `overlap` (called
    # `_overlap` until the original `overlap` is no longer needed)
    # is built from the old one.
    #
    # If the length of overlap exceeds the largest substring
    # seen so far (`sub_length`), we update the largest substring
    # to the overlapping strings.

    overlap = dict()
    # `sub_start_old` is the index of the beginning of the largest overlapping
    # substring in the old list. `sub_start_new` is the index of the beginning
    # of the same substring in the new list. `sub_length` is the length that
    # overlaps in both.
    # These track the largest overlapping substring seen so far, so naturally
    # we start with a 0-length substring.
    sub_start_old = 0
    sub_start_new = 0
    sub_length = 0

    for inew, val in enumerate(new):
        _overlap = dict()
        for iold in old_index_map.get(val,list()):
            # now we are considering all values of iold such that
            # `old[iold] == new[inew]`.
            _overlap[iold] = (iold and overlap.get(iold - 1, 0)) + 1
            if(_overlap[iold] > sub_length):
                # this is the largest substring seen so far, so store its
                # indices
                sub_length = _overlap[iold]
                sub_start_old = iold - sub_length + 1
                sub_start_new = inew - sub_length + 1
        overlap = _overlap

    if sub_length == 0:
        # If no common substring is found, we return an insert and delete...
        return (old and [('-', old)] or []) + (new and [('+', new)] or [])
    else:
        # ...otherwise, the common substring is unchanged and we recursively
        # diff the text before and after that substring
        return diff(old[ : sub_start_old], new[ : sub_start_new]) + \
               [('=', new[sub_start_new : sub_start_new + sub_length])] + \
               diff(old[sub_start_old + sub_length : ],
                       new[sub_start_new + sub_length : ])


def string_diff(old, new):
    '''
    Returns the difference between the old and new strings when split on
    whitespace. Considers punctuation a part of the word

    This function is intended as an example; you'll probably want
    a more sophisticated wrapper in practice.

    Params:
        old     the old string
        new     the new string

    Returns:
        the output of `diff` on the two strings after splitting them
        on whitespace (a list of change instructions; see the docstring
        of `diff`)

    Examples:
        >>> string_diff('The quick brown fox', 'The fast blue fox')
        ... # doctest: +NORMALIZE_WHITESPACE
        [('=', ['The']),
         ('-', ['quick', 'brown']),
         ('+', ['fast', 'blue']),
         ('=', ['fox'])]

    '''
    return diff(old.split(), new.split())


def html_diff(old, new):
    '''
    Returns the difference between two strings (as in stringDiff) in
    HTML format. HTML code in the strings is NOT escaped, so you
    will get weird results if the strings contain HTML.

    This function is intended as an example; you'll probably want
    a more sophisticated wrapper in practice.

    Params:
        old     the old string
        new     the new string

    Returns:
        the output of the diff expressed with HTML <ins> and <del>
        tags.

    Examples:
        >>> html_diff('The quick brown fox', 'The fast blue fox')
        'The <del>quick brown</del> <ins>fast blue</ins> fox'
    '''
    con = {'=': (lambda x: x),
           '+': (lambda x: "<ins>" + x + "</ins>"),
           '-': (lambda x: "<del>" + x + "</del>")}
    return " ".join([(con[a])(" ".join(b)) for a, b in string_diff(old, new)])


def check_diff(old, new):
    '''
    This tests that diffs returned by `diff` are valid. You probably won't
    want to use this function, but it's provided for documentation and
    testing.

    A diff should satisfy the property that the old input is equal to the
    elements of the result annotated with '-' or '=' concatenated together.
    Likewise, the new input is equal to the elements of the result annotated
    with '+' or '=' concatenated together. This function compares `old`,
    `new`, and the results of `diff(old, new)` to ensure this is true.

    Tests:
        >>> check_diff('ABCBA', 'CBABA')
        >>> check_diff('Foobarbaz', 'Foobarbaz')
        >>> check_diff('Foobarbaz', 'Boobazbam')
        >>> check_diff('The quick brown fox', 'Some quick brown car')
        >>> check_diff('A thick red book', 'A quick blue book')
        >>> check_diff('dafhjkdashfkhasfjsdafdasfsda', 'asdfaskjfhksahkfjsdha')
        >>> check_diff('88288822828828288282828', '88288882882828282882828')
        >>> check_diff('1234567890', '24689')
    '''
    old = list(old)
    new = list(new)
    result = diff(old, new)
    _old = [val for (a, vals) in result if (a in '=-') for val in vals]
    assert old == _old, 'Expected %s, got %s' % (old, _old)
    _new = [val for (a, vals) in result if (a in '=+') for val in vals]
    assert new == _new, 'Expected %s, got %s' % (new, _new)

############################################################################
############################## main ##################################
############################################################################

if __name__ == '__main__':
    main()
    exit(0)

